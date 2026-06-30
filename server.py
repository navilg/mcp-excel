import argparse
import base64
import io

from fastmcp import FastMCP
from openpyxl import load_workbook

mcp = FastMCP(name="Excel CRUD MCP")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _sanitize(value) -> str:
    """Replace newlines and tabs with literal \\n and \\t."""
    if value is None:
        return ""
    return str(value).replace("\n", "\\n").replace("\t", "\\t")


def _decode_excel(content: str) -> io.BytesIO:
    """Decode a base64-encoded Excel file into a BytesIO stream."""
    try:
        raw = base64.b64decode(content, validate=True)
    except Exception as exc:
        raise ValueError(
            "file_content must be a valid base64-encoded Excel file payload"
        ) from exc
    return io.BytesIO(raw)


def _encode_excel(wb) -> str:
    """Save a workbook to a base64-encoded string."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("ascii")


def _open_workbook(content: str, read_only: bool = False):
    """Decode and open a workbook from base64 content."""
    source = _decode_excel(content)
    return load_workbook(source, read_only=read_only)


def _get_sheet(wb, sheet_name: str | None = None):
    """Get a sheet by name or first sheet. Raises ValueError if not found."""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            )
        return wb[sheet_name]
    return wb[wb.sheetnames[0]]


def _sheet_to_tsv(ws) -> str:
    """Convert an openpyxl worksheet to a TSV string."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    lines = []
    for row in rows:
        lines.append("\t".join(_sanitize(cell) for cell in row))
    return "\n".join(lines)


def _get_column_idx(ws, column_name: str) -> int:
    """Return the 1-based column index for a column header name."""
    header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    for idx, cell in enumerate(header_row, start=1):
        if cell is not None and str(cell).strip() == column_name.strip():
            return idx
    raise ValueError(f"Column '{column_name}' not found in header row")


# =============================================
# BASE64 UTILITY tools
# =============================================


@mcp.tool
def encode_base64(data: str) -> str:
    """Encode plain text to a base64 string.

    Use this to encode a file's content (e.g. an Excel file read as text)
    into base64 so it can be passed to other tools like `read_excel`.

    Args:
        data: Plain text content to encode.
    """
    try:
        encoded = base64.b64encode(data.encode("utf-8")).decode("ascii")
        return encoded
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def decode_base64(encoded_data: str) -> str:
    """Decode a base64 string back to plain text.

    Use this to decode the base64 output returned by write tools
    (Create/Update/Delete) to verify the content, or to inspect
    the decoded payload of a base64-encoded Excel file.

    Args:
        encoded_data: Base64-encoded string to decode.
    """
    try:
        decoded = base64.b64decode(encoded_data, validate=True).decode("utf-8")
        return decoded
    except Exception as e:
        return f"Error: {e}"


# =============================================
# CREATE tools
# =============================================


@mcp.tool
def add_row(
    file_content: str,
    row_data: str,
    sheet_name: str | None = None,
) -> str:
    """Add a new row at the end of the sheet.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        row_data: Tab-separated values for the row (e.g. "val1\\tval2\\tval3").
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        values = row_data.split("\t")
        ws.append(values)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def add_cell_data(
    file_content: str,
    row_number: int,
    column_name: str,
    value: str,
    sheet_name: str | None = None,
) -> str:
    """Write a value into a specific cell identified by column name and row number.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        row_number: Row number (1 = first data row, header is row 1).
        column_name: Column header name.
        value: Value to write.
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        col_idx = _get_column_idx(ws, column_name)
        row_idx = row_number + 1  # +1 for header
        ws.cell(row=row_idx, column=col_idx, value=value)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def add_column(
    file_content: str,
    column_name: str,
    default_value: str = "",
    sheet_name: str | None = None,
) -> str:
    """Add a new column with a header name and optional default value for existing rows.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        column_name: Header name for the new column.
        default_value: Default value for existing rows (default: empty string).
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        new_col = (ws.max_column or 0) + 1
        # Write header
        ws.cell(row=1, column=new_col, value=column_name)
        # Fill default for existing data rows
        for r in range(2, (ws.max_row or 1) + 1):
            ws.cell(row=r, column=new_col, value=default_value)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def add_full_data(
    file_content: str,
    data: str,
    sheet_name: str | None = None,
) -> str:
    """Replace all data in a sheet with TSV-formatted content.

    The first line of `data` is treated as column headers.
    Subsequent lines are data rows.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        data: TSV-formatted data (header line + data lines, tab-separated).
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        # Clear existing content
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
        lines = data.strip().split("\n")
        for i, line in enumerate(lines):
            values = line.split("\t")
            for j, val in enumerate(values, start=1):
                ws.cell(row=i + 1, column=j, value=val)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def create_sheet(file_content: str, sheet_name: str) -> str:
    """Create a new sheet in the workbook.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        sheet_name: Name for the new sheet.
    """
    try:
        wb = _open_workbook(file_content)
        if sheet_name in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' already exists"
        wb.create_sheet(title=sheet_name)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


# =============================================
# READ tools
# =============================================


@mcp.tool
def read_excel(
    file_content: str,
    sheet_name: str | None = None,
) -> str:
    """Read an Excel sheet and return its content as TSV.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content, read_only=True)
        ws = _get_sheet(wb, sheet_name)
        result = _sheet_to_tsv(ws)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def read_rows_by_column(
    file_content: str,
    column_name: str,
    column_value: str,
    sheet_name: str | None = None,
) -> str:
    """Read rows where a specific column matches a value.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        column_name: Column header name to filter on.
        column_value: Value to match in the column.
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content, read_only=True)
        ws = _get_sheet(wb, sheet_name)
        col_idx = _get_column_idx(ws, column_name)

        headers = [
            cell
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        matching_rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_value = row[col_idx - 1]
            if (
                cell_value is not None
                and str(cell_value).strip() == column_value.strip()
            ):
                matching_rows.append(row)

        wb.close()

        if not matching_rows:
            return f"No rows found where '{column_name}' = '{column_value}'"

        result_lines = ["\t".join(_sanitize(h) for h in headers)]
        for row in matching_rows:
            result_lines.append("\t".join(_sanitize(c) for c in row))

        return "\n".join(result_lines)
    except Exception as e:
        return f"Error: {e}"


# =============================================
# UPDATE tools
# =============================================


@mcp.tool
def update_row(
    file_content: str,
    row_number: int,
    row_data: str,
    sheet_name: str | None = None,
) -> str:
    """Update an entire row with new values.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        row_number: Row number (1 = first data row, header is row 1).
        row_data: Tab-separated new values for the row.
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        values = row_data.split("\t")
        row_idx = row_number + 1  # +1 for header
        for j, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=j, value=val)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def update_cell(
    file_content: str,
    row_number: int,
    column_name: str,
    value: str,
    sheet_name: str | None = None,
) -> str:
    """Update a single cell by column name and row number.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        row_number: Row number (1 = first data row).
        column_name: Column header name.
        value: New value for the cell.
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        col_idx = _get_column_idx(ws, column_name)
        row_idx = row_number + 1
        ws.cell(row=row_idx, column=col_idx, value=value)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def replace_full_data(
    file_content: str,
    new_data: str,
    sheet_name: str | None = None,
) -> str:
    """Replace all data in a sheet with new TSV-formatted content.

    This is an alias for add_full_data with clearer semantics for updates.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        new_data: TSV-formatted data (first line = headers).
        sheet_name: Sheet name (defaults to first sheet).
    """
    return add_full_data(file_content, new_data, sheet_name)


# =============================================
# DELETE tools
# =============================================


@mcp.tool
def delete_sheet(file_content: str, sheet_name: str) -> str:
    """Delete a sheet from the workbook.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        sheet_name: Name of the sheet to delete.
    """
    try:
        wb = _open_workbook(file_content)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found"
        if len(wb.sheetnames) == 1:
            wb.close()
            return "Error: Cannot delete the only sheet in the workbook"
        del wb[sheet_name]
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def delete_row(
    file_content: str,
    row_number: int,
    sheet_name: str | None = None,
) -> str:
    """Delete a row from the sheet.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        row_number: Row number to delete (1 = first data row, header is row 1).
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        row_idx = row_number + 1  # +1 for header
        if ws.max_row and row_idx > ws.max_row:
            wb.close()
            return f"Error: Row {row_number} does not exist in sheet '{ws.title}'"
        ws.delete_rows(row_idx)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


@mcp.tool
def delete_column(
    file_content: str,
    column_name: str,
    sheet_name: str | None = None,
) -> str:
    """Delete a column by its header name from the sheet.

    Returns the modified workbook as base64 so the client can save it.

    Args:
        file_content: Base64-encoded Excel (.xlsx) file content.
        column_name: Column header name to delete.
        sheet_name: Sheet name (defaults to first sheet).
    """
    try:
        wb = _open_workbook(file_content)
        ws = _get_sheet(wb, sheet_name)
        col_idx = _get_column_idx(ws, column_name)
        ws.delete_cols(col_idx)
        result = _encode_excel(wb)
        wb.close()
        return result
    except Exception as e:
        return f"Error: {e}"


# =============================================
# Main
# =============================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel CRUD MCP Server")
    parser.add_argument(
        "--host",
        type=str,
        default="0.0.0.0",
        help="Host to bind the server to",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=8000,
        help="Port to bind the server to",
    )
    parser.add_argument(
        "--transport",
        type=str,
        default="streamable-http",
        help="Transport protocol (default: streamable-http)",
    )
    args = parser.parse_args()

    mcp.run(transport=args.transport, host=args.host, port=args.port)
